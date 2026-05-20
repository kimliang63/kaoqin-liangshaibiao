#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "output"

MAIN_SHEET = "每日打卡工时推送模版"
SIGN_SHEET = "美区签字报表"

# 默认取用户最新提供的数据目录；如你复制到项目内可改这里
INPUT_DIR = Path("/Users/masc/Downloads/0515排班")
MAIN_FILE = INPUT_DIR / "每日打卡工时推送模版.xlsx"
SIGN_FILE = INPUT_DIR / "美区签字报表 (13).xlsx"

REST_KEYWORDS = ("休息日", "节假日")
FLOW_KEYWORDS = ("休假", "公出", "出差")
WFH_KEYWORD = "居家办公"
FLOW_EXCLUDE_FOR_MISS = ("休假", "公出", "出差", "居家办公", "病假", "年假", "无薪")


def normalize_text(v):
    return "" if v is None else str(v).strip()


def to_float(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = normalize_text(v).replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(v))).date()
    s = normalize_text(v)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_datetime(v, base_day=None):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, time):
        d = base_day if base_day else date.today()
        return datetime.combine(d, v)
    if isinstance(v, (int, float)):
        x = float(v)
        if 0 <= x < 1:
            d = base_day if base_day else date.today()
            sec = int(round(x * 24 * 3600))
            return datetime.combine(d, time((sec // 3600) % 24, (sec % 3600) // 60, sec % 60))
        base = datetime(1899, 12, 30)
        return base + timedelta(days=x)
    s = normalize_text(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%H:%M:%S", "%H:%M"):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt in ("%H:%M:%S", "%H:%M"):
                d = base_day if base_day else date.today()
                return datetime.combine(d, dt.time())
            return dt
        except ValueError:
            continue
    return None


def safe_div(n, d):
    return None if d in (0, None) else n / d


def pct(v):
    return "N/A" if v is None else f"{v * 100:.2f}%"


def hours_between(start_dt, end_dt):
    if not start_dt or not end_dt:
        return 0.0
    if end_dt < start_dt:
        end_dt = end_dt + timedelta(days=1)
    return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)


def read_sheet_rows_auto(path, sheet_name, key_header="工号", max_scan_rows=10):
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"文件 {path} 缺少工作表：{sheet_name}")
    ws = wb[sheet_name]
    best = (0, 1, [])
    for r in range(1, max_scan_rows + 1):
        vals = [ws.cell(r, c).value for c in range(1, 120)]
        non = [v for v in vals if v not in (None, "")]
        score = len(non)
        if any(normalize_text(v) == key_header for v in non):
            score += 1000
        if score > best[0]:
            best = (score, r, vals)
    header_row = best[1]
    headers = []
    for v in best[2]:
        if v is None:
            if headers:
                break
            continue
        headers.append(normalize_text(v))
    rows = []
    empty_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        item = {}
        has_data = False
        for i, h in enumerate(headers, start=1):
            if not h:
                continue
            val = ws.cell(r, i).value
            item[h] = val
            if val not in (None, ""):
                has_data = True
        if has_data:
            rows.append(item)
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 200:
                break
    return rows


def make_hub(dept4, dept5):
    d4 = normalize_text(dept4)
    d5 = normalize_text(dept5)
    if ".H" in d5:
        return "HUB"
    if "Ground项目部" in d4 and d5 in {"EWR.G", "CNO.G"}:
        return "HUB"
    return "NON_HUB"


def is_rest_shift(shift_name):
    s = normalize_text(shift_name)
    return any(k in s for k in REST_KEYWORDS)


def has_any_flow(remark):
    r = normalize_text(remark)
    return any(k in r for k in FLOW_EXCLUDE_FOR_MISS)


def classify_group(dept3, dept4):
    d3 = normalize_text(dept3)
    d4 = normalize_text(dept4)
    if "大区" in d3:
        return d3, d4 if d4 else "未分组"
    return "中后台", d3 if d3 else "未分组"


def schedule_correct_flag(shift_name, shift_start, shift_end, first_card, last_card, remark, wfh_hours):
    shift = normalize_text(shift_name)
    rem = normalize_text(remark)
    if not shift:
        return "未排班"

    any_card = first_card is not None or last_card is not None
    complete_cards = first_card is not None and last_card is not None

    if is_rest_shift(shift):
        if not any_card:
            return "正确"
        if not complete_cards:
            return "不正确"
        if not shift_start:
            return "不正确"
        diff_h = abs((first_card - shift_start).total_seconds()) / 3600.0
        return "正确" if diff_h <= 1 else "不正确"

    if any(k in rem for k in FLOW_KEYWORDS):
        return "正确"

    if WFH_KEYWORD in rem:
        if not (complete_cards and shift_start and shift_end):
            return "不正确"
        covered = hours_between(first_card, last_card) + wfh_hours
        shift_h = hours_between(shift_start, shift_end)
        return "正确" if covered >= shift_h else "不正确"

    if any_card and not complete_cards:
        return "不正确"
    if not complete_cards:
        return "不正确"
    if not shift_start:
        return "不正确"
    diff_h = abs((first_card - shift_start).total_seconds()) / 3600.0
    return "正确" if diff_h <= 1 else "不正确"


def build_sign_map(sign_rows):
    sign_map = {}
    for r in sign_rows:
        emp = normalize_text(r.get("工号"))
        d = to_date(r.get("考勤日期"))
        if not emp or not d:
            continue
        key = f"{emp}_{d.isoformat()}"
        cumulative = None
        for c in ("累计总工时", "累计总工时（GF）", "累计总工时（签字）", "计薪出勤时长合计（REG)"):
            if c in r and r.get(c) not in (None, ""):
                cumulative = to_float(r.get(c))
                break
        sign_map[key] = {
            "累计总工时": cumulative,
            "签字核实": normalize_text(r.get("员工签字核实（GF）")),
        }
    return sign_map


def build_records(main_rows, sign_map):
    records = []
    for row in main_rows:
        emp = normalize_text(row.get("工号"))
        if not emp:
            continue
        day = to_date(row.get("考勤日期"))
        if not day:
            continue
        key = f"{emp}_{day.isoformat()}"
        dept3 = normalize_text(row.get("三级部门"))
        dept4 = normalize_text(row.get("四级部门"))
        dept5 = normalize_text(row.get("五级部门"))
        shift_name = normalize_text(row.get("班次名称"))
        remark = normalize_text(row.get("备注（GF）"))

        shift_start = to_datetime(row.get("班次上班时间"), day)
        shift_end = to_datetime(row.get("班次下班时间"), day)
        first_card = to_datetime(row.get("首打卡时间"), day)
        last_card = to_datetime(row.get("末打卡时间"), day)

        wfh_full = to_float(row.get("居家办公工时（全）合计"))
        wfh_pending = to_float(row.get("居家办公合计（审批中）"))
        wfh_total = wfh_full + wfh_pending

        daily_total_src = to_float(row.get("每日总工时(公式：末打卡-首打卡-班次午休时间+居家办公时长)合计"))
        daily_total = daily_total_src + wfh_pending
        over8_extra = daily_total - 8
        over8_flag = "是" if over8_extra > 0 else "否"

        sign = sign_map.get(key, {})
        cumulative_main = to_float(row.get("累计总工时"))
        cumulative_sign = sign.get("累计总工时")
        cumulative_total = cumulative_sign if cumulative_sign is not None else cumulative_main
        week_ot = max(cumulative_total - 40, 0)

        cards_count = int(round(to_float(row.get("班次内打卡次数"))))
        should_count_cards = bool(shift_name) and not is_rest_shift(shift_name) and not has_any_flow(remark)
        miss_cards = max(4 - cards_count, 0) if should_count_cards else 0

        schedule_flag = "是" if shift_name else "否"
        correct_flag = schedule_correct_flag(shift_name, shift_start, shift_end, first_card, last_card, remark, wfh_total)

        region, dept = classify_group(dept3, dept4)
        hub = make_hub(dept4, dept5)

        records.append(
            {
                "考勤日期": day.isoformat(),
                "区域": region,
                "部门": dept,
                "姓名": normalize_text(row.get("姓名")),
                "工号": emp,
                "三级部门": dept3,
                "四级部门": dept4,
                "五级部门": dept5,
                "hub分类": hub,
                "班次名称": shift_name,
                "班次上班时间": shift_start.strftime("%H:%M") if shift_start else "",
                "班次下班时间": shift_end.strftime("%H:%M") if shift_end else "",
                "首打卡时间": first_card.strftime("%H:%M") if first_card else "",
                "末打卡时间": last_card.strftime("%H:%M") if last_card else "",
                "备注（GF）": remark,
                "是否排班": schedule_flag,
                "排班正确": correct_flag,
                "排班正确计数": 1 if correct_flag == "正确" else 0,
                "日超8H人数计数": 1 if over8_flag == "是" else 0,
                "是否日超8H": over8_flag,
                "日超8H时长": round(max(over8_extra, 0), 2),
                "居家办公合计（审批中）": round(wfh_pending, 2),
                "每日总工时": round(daily_total, 2),
                "累计总工时": round(cumulative_total, 2),
                "本周加班工时": round(week_ot, 2),
                "班次内打卡次数": cards_count,
                "缺卡次数": miss_cards,
                "应打卡人数计数": 1 if should_count_cards else 0,
                "签字核实": normalize_text(sign.get("签字核实")),
            }
        )
    return records


def build_day_group_summary(records):
    by_day_group = defaultdict(
        lambda: {
            "总人数": 0,
            "日超8h人数": 0,
            "未排班数": 0,
            "排班人数": 0,
            "HUB人数": 0,
            "HUB排班正确人数": 0,
            "缺卡次数": 0,
            "应打卡人数": 0,
            "排班正确人数": 0,
        }
    )
    for r in records:
        key = (r["考勤日期"], r["区域"], r["部门"])
        g = by_day_group[key]
        g["总人数"] += 1
        g["日超8h人数"] += r["日超8H人数计数"]
        if r["是否排班"] == "否":
            g["未排班数"] += 1
        else:
            g["排班人数"] += 1
        if r["hub分类"] == "HUB":
            g["HUB人数"] += 1
            if r["排班正确"] == "正确":
                g["HUB排班正确人数"] += 1
        g["缺卡次数"] += r["缺卡次数"]
        g["应打卡人数"] += r["应打卡人数计数"]
        g["排班正确人数"] += r["排班正确计数"]
    return by_day_group


def summarize(records):
    if not records:
        raise ValueError("主数据为空，无法生成报表。")
    days = sorted({r["考勤日期"] for r in records})
    current_day = days[-1]
    prev_day = days[-2] if len(days) >= 2 else None

    by_day_group = build_day_group_summary(records)

    current_rows = [r for r in records if r["考勤日期"] == current_day]
    total = len(current_rows)
    scheduled = sum(1 for r in current_rows if r["是否排班"] == "是")
    unscheduled = total - scheduled
    correct = sum(r["排班正确计数"] for r in current_rows)
    over8 = sum(r["日超8H人数计数"] for r in current_rows)
    miss_sum = sum(r["缺卡次数"] for r in current_rows)
    should_people = sum(r["应打卡人数计数"] for r in current_rows)
    miss_rate = safe_div(miss_sum, should_people * 4 if should_people else 0)
    hub_total = sum(1 for r in current_rows if r["hub分类"] == "HUB")
    hub_correct = sum(1 for r in current_rows if r["hub分类"] == "HUB" and r["排班正确"] == "正确")

    kpi = {
        "总人数": total,
        "日超8h人数": over8,
        "未排班数": unscheduled,
        "排班率": safe_div(scheduled, total),
        "HUB排班正确率": safe_div(hub_correct, hub_total),
        "排班正确率": safe_div(correct, total),
        "缺卡数": miss_sum,
        "缺卡率": miss_rate,
    }

    kpi_trend = {}
    if prev_day:
        prev_rows = [r for r in records if r["考勤日期"] == prev_day]
        p_total = len(prev_rows)
        p_scheduled = sum(1 for r in prev_rows if r["是否排班"] == "是")
        p_correct = sum(r["排班正确计数"] for r in prev_rows)
        p_over8 = sum(r["日超8H人数计数"] for r in prev_rows)
        p_miss_sum = sum(r["缺卡次数"] for r in prev_rows)
        p_should = sum(r["应打卡人数计数"] for r in prev_rows)
        p_hub_total = sum(1 for r in prev_rows if r["hub分类"] == "HUB")
        p_hub_correct = sum(1 for r in prev_rows if r["hub分类"] == "HUB" and r["排班正确"] == "正确")
        prev_kpi = {
            "排班率": safe_div(p_scheduled, p_total),
            "HUB排班正确率": safe_div(p_hub_correct, p_hub_total),
            "排班正确率": safe_div(p_correct, p_total),
            "缺卡率": safe_div(p_miss_sum, p_should * 4 if p_should else 0),
            "日超8h人数": p_over8,
        }
        for key in ("排班率", "HUB排班正确率", "排班正确率", "缺卡率"):
            kpi_trend[key] = safe_div((kpi[key] or 0) - (prev_kpi[key] or 0), prev_kpi[key] or 0)
        kpi_trend["日超8h人数"] = safe_div((kpi["日超8h人数"] or 0) - (prev_kpi["日超8h人数"] or 0), prev_kpi["日超8h人数"] or 0)
        kpi_trend["对比周期"] = f"{prev_day} -> {current_day}"

    prev_map = {}
    if prev_day:
        for (d, region, dept), val in by_day_group.items():
            if d == prev_day:
                prev_map[(region, dept)] = val

    group_rows = []
    for (d, region, dept), val in by_day_group.items():
        if d != current_day:
            continue
        prev_val = prev_map.get((region, dept), {})
        row = {
            "区域": region,
            "部门": dept,
            "总人数": val["总人数"],
            "日超8h人": val["日超8h人数"],
            "未排班数": val["未排班数"],
            "排班率": safe_div(val["排班人数"], val["总人数"]),
            "HUB排班正确": safe_div(val["HUB排班正确人数"], val["HUB人数"]),
            "缺卡数": val["缺卡次数"],
            "缺卡率": safe_div(val["缺卡次数"], val["应打卡人数"] * 4 if val["应打卡人数"] else 0),
            "前一日排班率": None,
            "前一日缺卡率": None,
            "排班率变化": None,
            "缺卡率变化": None,
            "排班正确率": safe_div(val["排班正确人数"], val["总人数"]),
        }
        if prev_val:
            p_schedule = safe_div(prev_val["排班人数"], prev_val["总人数"])
            p_miss = safe_div(prev_val["缺卡次数"], prev_val["应打卡人数"] * 4 if prev_val["应打卡人数"] else 0)
            row["前一日排班率"] = p_schedule
            row["前一日缺卡率"] = p_miss
            row["排班率变化"] = safe_div((row["排班率"] or 0) - (p_schedule or 0), p_schedule or 0)
            row["缺卡率变化"] = safe_div((row["缺卡率"] or 0) - (p_miss or 0), p_miss or 0)
        group_rows.append(row)

    group_rows.sort(key=lambda x: (0 if "中后台" not in x["区域"] else 1, -(x["总人数"])))

    details = sorted(current_rows, key=lambda x: (x["区域"], x["部门"], x["姓名"]))
    return {
        "meta": {
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "当前数据日": current_day,
            "前一日": prev_day or "",
            "数据源目录": str(INPUT_DIR),
        },
        "kpi": kpi,
        "kpi_trend": kpi_trend,
        "group_rows": group_rows,
        "detail_rows": details,
    }


def to_js(data):
    return json.dumps(data, ensure_ascii=False)


def build_html(data):
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>GUS_排班情况总览</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "PingFang SC", "Microsoft YaHei", sans-serif; }}
    .table-wrap {{ overflow-x: auto; border: 1px solid #d1d5db; border-radius: 10px; background: #fff; }}
    table {{ border-collapse: collapse; min-width: 1200px; width: 100%; }}
    th, td {{ border: 1px solid #d1d5db; padding: 6px 8px; font-size: 12px; text-align: center; white-space: nowrap; }}
    th {{ background: #e2e8f0; font-weight: 600; }}
    .good {{ background: #8bc34a; color: #111827; font-weight: 600; }}
    .warn {{ background: #f6d365; color: #111827; font-weight: 600; }}
    .bad {{ background: #ef5350; color: #ffffff; font-weight: 600; }}
    .delta-up {{ color: #16a34a; font-weight: 700; }}
    .delta-down {{ color: #dc2626; font-weight: 700; }}
    .delta-flat {{ color: #64748b; font-weight: 700; }}
    .kpi-card {{ border:1px solid #cbd5e1;border-radius:10px;padding:10px;background:#fff; }}
  </style>
</head>
<body class="bg-slate-50 text-slate-900">
  <div class="max-w-[1700px] mx-auto p-4">
    <h1 class="text-2xl font-bold">GUS_排班情况总览</h1>
    <p class="text-sm text-slate-600 mt-1">数据日：<span id="curDay"></span> ｜ 生成时间：<span id="genAt"></span></p>

    <section class="mt-4">
      <h2 class="text-lg font-semibold mb-2">▼ 排班整体情况汇总</h2>
      <div class="grid grid-cols-2 md:grid-cols-4 xl:grid-cols-8 gap-2" id="kpis"></div>
    </section>

    <section class="mt-4 table-wrap p-2">
      <details open>
        <summary class="font-semibold cursor-pointer">▼ 区域表现（重点）</summary>
        <div class="flex gap-2 my-2">
          <input id="kw" class="border rounded px-2 py-1 text-sm" placeholder="搜索区域/部门" />
          <select id="type" class="border rounded px-2 py-1 text-sm">
            <option value="">全部</option>
            <option value="大区">仅大区</option>
            <option value="中后台">仅中后台</option>
          </select>
          <button id="applyRegion" class="bg-slate-900 text-white rounded px-3 py-1 text-sm">筛选</button>
        </div>
        <div id="summaryTable"></div>
      </details>
    </section>

    <section class="mt-4 table-wrap p-2">
      <details open>
        <summary class="font-semibold cursor-pointer">▼ 排班及考勤明细</summary>
        <div class="flex gap-2 my-2">
          <input id="dkw" class="border rounded px-2 py-1 text-sm" placeholder="搜索姓名/工号/部门" />
          <select id="dPageSize" class="border rounded px-2 py-1 text-sm">
            <option value="20">每页20条</option>
            <option value="50">每页50条</option>
            <option value="100">每页100条</option>
          </select>
          <button id="applyDetail" class="bg-slate-900 text-white rounded px-3 py-1 text-sm">筛选</button>
        </div>
        <div id="detailTable"></div>
        <div class="flex justify-between items-center mt-2">
          <div id="pageInfo" class="text-sm text-slate-600"></div>
          <div class="flex gap-2">
            <button id="prevPage" class="border rounded px-3 py-1 text-sm">上一页</button>
            <button id="nextPage" class="border rounded px-3 py-1 text-sm">下一页</button>
          </div>
        </div>
      </details>
    </section>

    <section class="mt-4 table-wrap p-2">
      <h2 class="font-semibold">数据口径</h2>
      <ul class="list-disc pl-5 text-sm leading-6">
        <li>班次名称为空标记未排班；有排班时，首打卡与班次上班时间差值在±1小时内为排班正确。</li>
        <li>节假日/休息日无首末卡标记正确；有打卡则按正常班次规则判断；有打卡但首末卡不完整标记不正确。</li>
        <li>工作日备注含休假/公出/出差流程直接标记正确；含居家办公按“首末卡时长+居家工时是否覆盖班次时长”判断。</li>
        <li>日超8h人数：`每日总工时 + 居家审批中 - 8 > 0` 的人数。</li>
        <li>缺卡率：`缺卡次数之和 / (应打卡人数 × 4)`；休息日与提交流程人员不计入应打卡人数。</li>
      </ul>
    </section>
  </div>
  <script>
    const DATA = {to_js(data)};
    const pct = (v) => v == null ? "/" : (v * 100).toFixed(0) + "%";
    const pct2 = (v) => v == null ? "/" : (v * 100).toFixed(2) + "%";
    const trend = (v) => {{
      if (v == null || !isFinite(v)) return '<span class="delta-flat">--</span>';
      if (v > 0) return '<span class="delta-up">↑ ' + (v * 100).toFixed(1) + '%</span>';
      if (v < 0) return '<span class="delta-down">↓ ' + Math.abs(v * 100).toFixed(1) + '%</span>';
      return '<span class="delta-flat">→ 0.0%</span>';
    }};
    const colorSchedule = (v) => {{
      if (v == null) return '';
      if (v >= 1) return 'good';
      if (v >= 0.9) return 'warn';
      return 'bad';
    }};
    const colorCorrect = (v) => {{
      if (v == null) return '';
      if (v >= 0.9) return 'good';
      if (v >= 0.5) return 'warn';
      return 'bad';
    }};
    const colorMiss = (v) => {{
      if (v == null) return '';
      if (v >= 0.9) return 'bad';
      if (v >= 0.5) return 'warn';
      return 'good';
    }};

    document.getElementById('curDay').textContent = DATA.meta['当前数据日'];
    document.getElementById('genAt').textContent = DATA.meta['生成时间'];

    const kpi = DATA.kpi || {{}};
    const kpiTrend = DATA.kpi_trend || {{}};
    const kpiItems = [
      ['总人数', kpi['总人数'], null],
      ['日超8h人数', kpi['日超8h人数'], '日超8h人数'],
      ['未排班数', kpi['未排班数'], null],
      ['排班率', pct2(kpi['排班率']), '排班率'],
      ['HUB排班正确率', pct2(kpi['HUB排班正确率']), 'HUB排班正确率'],
      ['排班正确率', pct2(kpi['排班正确率']), '排班正确率'],
      ['缺卡数', kpi['缺卡数'], null],
      ['缺卡率', pct2(kpi['缺卡率']), '缺卡率'],
    ];
    document.getElementById('kpis').innerHTML = kpiItems.map(([k, v, tk]) =>
      `<div class="kpi-card"><div class="text-xs text-slate-500">${{k}}</div><div class="text-xl font-semibold">${{v ?? '/'}}</div><div class="text-xs">${{tk ? trend(kpiTrend[tk]) : '--'}}</div></div>`
    ).join('');

    const summarySource = DATA.group_rows || [];
    function renderSummary(rows) {{
      let html = '<table><thead><tr>' +
        '<th>区域</th><th>部门</th><th>总人数</th><th>日超8h人</th><th>未排班数</th>' +
        `<th>排班率</th><th>HUB排班正确</th><th>缺卡数</th><th>缺卡率</th>` +
        `<th>${{DATA.meta['前一日'] || '前一日'}}排班率</th><th>${{DATA.meta['前一日'] || '前一日'}}缺卡率</th>` +
        '<th>排班率变化</th><th>缺卡率变化</th></tr></thead><tbody>';
      for (const r of rows) {{
        html += '<tr>' +
          `<td>${{r['区域']}}</td>` +
          `<td>${{r['部门']}}</td>` +
          `<td>${{r['总人数']}}</td>` +
          `<td>${{r['日超8h人']}}</td>` +
          `<td>${{r['未排班数']}}</td>` +
          `<td class="${{colorSchedule(r['排班率'])}}">${{pct(r['排班率'])}}</td>` +
          `<td class="${{colorCorrect(r['HUB排班正确'])}}">${{pct(r['HUB排班正确'])}}</td>` +
          `<td>${{r['缺卡数']}}</td>` +
          `<td class="${{colorMiss(r['缺卡率'])}}">${{pct(r['缺卡率'])}}</td>` +
          `<td>${{pct(r['前一日排班率'])}}</td>` +
          `<td>${{pct(r['前一日缺卡率'])}}</td>` +
          `<td>${{trend(r['排班率变化'])}}</td>` +
          `<td>${{trend(r['缺卡率变化'])}}</td>` +
          '</tr>';
      }}
      html += '</tbody></table>';
      document.getElementById('summaryTable').innerHTML = html;
    }}
    renderSummary(summarySource);
    document.getElementById('applyRegion').addEventListener('click', () => {{
      const kw = (document.getElementById('kw').value || '').trim();
      const tp = document.getElementById('type').value;
      const rows = summarySource.filter(r => {{
        const txt = (r['区域'] + ' ' + r['部门']);
        if (kw && !txt.includes(kw)) return false;
        if (tp === '大区' && r['区域'] === '中后台') return false;
        if (tp === '中后台' && r['区域'] !== '中后台') return false;
        return true;
      }});
      renderSummary(rows);
    }});

    const detailSource = DATA.detail_rows || [];
    let detailFiltered = detailSource.slice();
    let page = 1;
    function renderDetail() {{
      const pageSize = parseInt(document.getElementById('dPageSize').value, 10);
      const pages = Math.max(1, Math.ceil(detailFiltered.length / pageSize));
      if (page > pages) page = pages;
      if (page < 1) page = 1;
      const start = (page - 1) * pageSize;
      const rows = detailFiltered.slice(start, start + pageSize);
      let html = '<table><thead><tr>' +
        '<th>区域</th><th>部门</th><th>姓名</th><th>工号</th><th>班次名称</th><th>首打卡</th><th>末打卡</th>' +
        '<th>是否排班</th><th>排班正确</th><th>日超8h</th><th>累计总工时</th><th>本周加班工时</th><th>缺卡次数</th><th>备注</th></tr></thead><tbody>';
      for (const r of rows) {{
        html += '<tr>' +
          `<td>${{r['区域']}}</td><td>${{r['部门']}}</td><td>${{r['姓名']}}</td><td>${{r['工号']}}</td>` +
          `<td>${{r['班次名称'] || '/'}}</td><td>${{r['首打卡时间'] || '/'}}</td><td>${{r['末打卡时间'] || '/'}}</td>` +
          `<td>${{r['是否排班']}}</td>` +
          `<td class="${{r['排班正确']==='正确'?'good':(r['排班正确']==='不正确'?'bad':'')}}">${{r['排班正确']}}</td>` +
          `<td>${{r['是否日超8H']}}</td><td>${{r['累计总工时']}}</td><td>${{r['本周加班工时']}}</td>` +
          `<td>${{r['缺卡次数']}}</td><td>${{r['备注（GF）'] || '/'}}</td>` +
          '</tr>';
      }}
      html += '</tbody></table>';
      document.getElementById('detailTable').innerHTML = html;
      document.getElementById('pageInfo').textContent = `第${{page}}/${{pages}}页，共${{detailFiltered.length}}条`;
    }}
    document.getElementById('applyDetail').addEventListener('click', () => {{
      const kw = (document.getElementById('dkw').value || '').trim();
      detailFiltered = detailSource.filter(r => {{
        if (!kw) return true;
        return (r['姓名'] + ' ' + r['工号'] + ' ' + r['区域'] + ' ' + r['部门']).includes(kw);
      }});
      page = 1;
      renderDetail();
    }});
    document.getElementById('dPageSize').addEventListener('change', () => {{ page = 1; renderDetail(); }});
    document.getElementById('prevPage').addEventListener('click', () => {{ page -= 1; renderDetail(); }});
    document.getElementById('nextPage').addEventListener('click', () => {{ page += 1; renderDetail(); }});
    renderDetail();
  </script>
</body>
</html>
"""


def generate_report_data(main_file, sign_file):
    main_rows = read_sheet_rows_auto(main_file, MAIN_SHEET, key_header="工号")
    sign_rows = read_sheet_rows_auto(sign_file, SIGN_SHEET, key_header="工号")
    sign_map = build_sign_map(sign_rows)
    records = build_records(main_rows, sign_map)
    result = summarize(records)
    html = build_html(result)
    return result, html


def generate_and_write_outputs(main_file, sign_file, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_json = output_dir / "attendance_report_data.json"
    out_html = output_dir / "attendance_report.html"
    result, html = generate_report_data(main_file, sign_file)
    out_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    out_html.write_text(html, encoding="utf-8")
    return out_json, out_html, result


def main():
    out_json, out_html, _ = generate_and_write_outputs(
        main_file=MAIN_FILE,
        sign_file=SIGN_FILE,
        output_dir=OUTPUT_DIR,
    )
    print(f"已生成: {out_html}")
    print(f"已生成: {out_json}")


if __name__ == "__main__":
    main()

