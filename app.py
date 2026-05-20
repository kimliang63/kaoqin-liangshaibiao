#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import tempfile
import csv
import io
from pathlib import Path
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook

from generate_report import (
    MAIN_SHEET,
    SIGN_SHEET,
    generate_report_data,
)


MAX_FILE_MB = 50
BASE_DIR = Path(__file__).resolve().parent
HISTORY_DIR = BASE_DIR / "output" / "history"


def _check_file_size(uploaded_file, max_mb=MAX_FILE_MB):
    size_mb = uploaded_file.size / (1024 * 1024)
    if size_mb > max_mb:
        raise ValueError(f"{uploaded_file.name} 超过 {max_mb}MB，当前 {size_mb:.1f}MB。")


def _check_required_sheet(path, expected_sheet, file_label):
    wb = load_workbook(path, read_only=True, data_only=True)
    if expected_sheet not in wb.sheetnames:
        raise ValueError(f"{file_label} 缺少工作表：{expected_sheet}")


def _save_uploaded(uploaded_file, target_path):
    target_path.write_bytes(uploaded_file.getbuffer())


def _trend_text(v):
    if v is None:
        return "--"
    if v > 0:
        return f"↑ {v * 100:.2f}%"
    if v < 0:
        return f"↓ {abs(v * 100):.2f}%"
    return "→ 0.00%"


def _safe_date_text(s):
    return (s or "").replace("/", "-").replace(" ", "_")


def _history_paths(report_date):
    d = HISTORY_DIR / _safe_date_text(report_date)
    return {
        "dir": d,
        "json": d / "attendance_report_data.json",
        "html": d / "attendance_report.html",
        "csv": d / "attendance_report_detail.csv",
    }


def _save_history(result, html, json_text):
    report_date = result.get("meta", {}).get("当前数据日")
    if not report_date:
        return
    p = _history_paths(report_date)
    p["dir"].mkdir(parents=True, exist_ok=True)
    p["json"].write_text(json_text, encoding="utf-8")
    p["html"].write_text(html, encoding="utf-8")

    detail_rows = result.get("detail_rows", [])
    if detail_rows:
        csv_buf = io.StringIO()
        writer = csv.DictWriter(csv_buf, fieldnames=list(detail_rows[0].keys()))
        writer.writeheader()
        writer.writerows(detail_rows)
        p["csv"].write_text(csv_buf.getvalue(), encoding="utf-8-sig")


def _load_history_reports():
    reports = {}
    if not HISTORY_DIR.exists():
        return reports
    for d in HISTORY_DIR.iterdir():
        if not d.is_dir():
            continue
        json_path = d / "attendance_report_data.json"
        html_path = d / "attendance_report.html"
        if not (json_path.exists() and html_path.exists()):
            continue
        try:
            result = json.loads(json_path.read_text(encoding="utf-8"))
            report_date = result.get("meta", {}).get("当前数据日") or d.name
            reports[report_date] = {
                "result": result,
                "html": html_path.read_text(encoding="utf-8"),
                "json_text": json.dumps(result, ensure_ascii=False, indent=2),
            }
        except Exception:
            continue
    return reports


def _kpi_delta_text(cur, prev, is_percent=False):
    if cur is None or prev is None:
        return "--"
    diff = cur - prev
    if is_percent:
        return f"{diff * 100:+.2f}%"
    return f"{diff:+.0f}"


st.set_page_config(page_title="GUS_排班情况总览", layout="wide")
st.title("GUS_排班情况总览（公网交互版）")
st.caption("上传数据 -> 一键生成考勤晾晒结果 -> 自动保存历史 -> 按日期切换查看与跨次对比")

with st.expander("上传说明", expanded=False):
    st.markdown(
        "- 必传：每日打卡工时推送模版.xlsx + 签字报表.xlsx\n"
        "- 系统将自动解析 xlsx，生成结构化 JSON 与考勤报表\n"
        f"- 单文件上限：{MAX_FILE_MB}MB"
    )

col1, col2 = st.columns(2)
with col1:
    main_file = st.file_uploader("1) 每日打卡工时推送模版（xlsx）", type=["xlsx"], key="main_file")
with col2:
    sign_file = st.file_uploader("2) 签字报表（xlsx）", type=["xlsx"], key="sign_file")

if st.button("生成报表", type="primary", use_container_width=True):
    if not all([main_file, sign_file]):
        st.error("请先上传2个必传文件后再生成。")
        st.stop()

    try:
        for f in [main_file, sign_file]:
            _check_file_size(f)

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            main_path = td_path / "main.xlsx"
            sign_path = td_path / "sign.xlsx"
            _save_uploaded(main_file, main_path)
            _save_uploaded(sign_file, sign_path)

            # 输入校验：必须工作表
            _check_required_sheet(main_path, MAIN_SHEET, "主明细文件")
            _check_required_sheet(sign_path, SIGN_SHEET, "签字报表文件")

            with st.spinner("正在生成，请稍候..."):
                result, html = generate_report_data(
                    main_file=main_path,
                    sign_file=sign_path,
                )

        st.success("生成成功。")
        st.session_state["report_result"] = result
        st.session_state["report_html"] = html
        st.session_state["report_json_text"] = json.dumps(result, ensure_ascii=False, indent=2)
        _save_history(
            result=st.session_state["report_result"],
            html=st.session_state["report_html"],
            json_text=st.session_state["report_json_text"],
        )

    except Exception as e:
        st.error(f"生成失败：{e}")

history_reports = _load_history_reports()
available_dates = sorted(history_reports.keys(), reverse=True)

if "report_result" in st.session_state and "report_html" in st.session_state:
    report_date = st.session_state["report_result"].get("meta", {}).get("当前数据日")
    if report_date and report_date not in history_reports:
        history_reports[report_date] = {
            "result": st.session_state["report_result"],
            "html": st.session_state["report_html"],
            "json_text": st.session_state["report_json_text"],
        }
        available_dates = sorted(history_reports.keys(), reverse=True)

if available_dates:
    st.subheader("历史数据切换")
    cdate1, cdate2 = st.columns(2)
    with cdate1:
        selected_date = st.selectbox("查看日期", options=available_dates, index=0)
    compare_options = ["不对比"] + [d for d in available_dates if d != selected_date]
    with cdate2:
        compare_date = st.selectbox("对比日期", options=compare_options, index=0)

    selected_pack = history_reports[selected_date]
    result = selected_pack["result"]
    html = selected_pack["html"]
    json_text = selected_pack["json_text"]
    kpi = result.get("kpi", {})
    compare_kpi = history_reports.get(compare_date, {}).get("result", {}).get("kpi", {}) if compare_date != "不对比" else {}
    detail_rows = result.get("detail_rows", [])

    st.subheader("核心指标")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("总人数", int(kpi.get("总人数", 0)), _kpi_delta_text(kpi.get("总人数"), compare_kpi.get("总人数")))
    c2.metric("未排班数", int(kpi.get("未排班数", 0)), _kpi_delta_text(kpi.get("未排班数"), compare_kpi.get("未排班数")))
    c3.metric("排班率", f"{(kpi.get('排班率') or 0)*100:.2f}%", _kpi_delta_text(kpi.get("排班率"), compare_kpi.get("排班率"), True))
    c4.metric(
        "HUB排班正确率",
        f"{(kpi.get('HUB排班正确率') or 0)*100:.2f}%",
        _kpi_delta_text(kpi.get("HUB排班正确率"), compare_kpi.get("HUB排班正确率"), True),
    )
    c5.metric("缺卡率", f"{(kpi.get('缺卡率') or 0)*100:.2f}%", _kpi_delta_text(kpi.get("缺卡率"), compare_kpi.get("缺卡率"), True))

    if compare_date != "不对比":
        st.caption(f"当前查看：{selected_date}；对比基准：{compare_date}")
    else:
        st.caption(f"当前查看：{selected_date}；未选择对比日期")

    st.subheader("下载结果")
    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "下载 attendance_report.html",
        data=html.encode("utf-8"),
        file_name="attendance_report.html",
        mime="text/html",
        use_container_width=True,
    )
    d2.download_button(
        "下载 attendance_report_data.json",
        data=json_text.encode("utf-8"),
        file_name="attendance_report_data.json",
        mime="application/json",
        use_container_width=True,
    )
    if detail_rows:
        csv_buf = io.StringIO()
        writer = csv.DictWriter(csv_buf, fieldnames=list(detail_rows[0].keys()))
        writer.writeheader()
        writer.writerows(detail_rows)
        d3.download_button(
            "下载 attendance_report_detail.csv",
            data=csv_buf.getvalue().encode("utf-8-sig"),
            file_name="attendance_report_detail.csv",
            mime="text/csv",
            use_container_width=True,
        )

    st.subheader("在线预览报表")
    st.components.v1.html(html, height=1200, scrolling=True)
